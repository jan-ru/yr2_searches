"""This module converts an excel table to a markdown table. Numbers show a
    thousands separator and will be right aligned. A line separator before
    the total can be inserted.

The excel table is loaded into a dataframe. The thousands separator is
set to dot. The last line of the dataframe is duplicated. In the line before
last, a separator-line is inserted.

Args:
    file_name (str): the filename of the excel data file
    sheet_name (str): the sheetname in the excel file to load the data from
    table_name (str): the tablename in the excel sheet to load the data from

Example:
    xls2md("sia-projecten.xlsx", "ThemaSelect", "Thema")

Attributes:
    df is the Pandas DataFrame that the excel table is loaded into

To do:
     * generalize the module such that it accepts tables with multiple
        numeric columns maybe later pass a list of dictionary's where
        the dictionary holds {column name : sum=true} or {column name
        : sum = false}
     * include a test check_table_exists -> True/False
     * possibly use argparse as shown in excel2markdown by SpeerSec

.. _Ref excel2markdown by SpeerSec:
   https://github.com/SpeerSec/excel2markdown
"""

import sys
import locale
import pandas as pd
from pandas.api.types import is_numeric_dtype
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_to_tuple


def xls2md(table: dict) -> pd.DataFrame:
    """
    Args:
        table

    """

    locale.setlocale(locale.LC_ALL, "de_DE")
    # print(table["sheet"] + "!" + table["range"])
    # note that an error may occur if the sheetname is not one word
    # The test if the table exists is (now) done prior to calling this function
    coordinaten: list = unpack_xy(table["sheet"] + "!" + table["range"])
    xls_tables = pd.ExcelFile(table["file"])
    with xls_tables as xls:
        df = pd.read_excel(
            xls,
            table["sheet"],
            skiprows=coordinaten[1] - 1,  # start reading
            nrows=coordinaten[3] - coordinaten[1],  # rows to read
            usecols=coordinaten[0] + ":" + coordinaten[2],  # cols to read
        )

    # err: total is created even if there is not total row
    if is_numeric_dtype(df[df.columns[1]]):
        df[df.columns[1]] = df[df.columns[1]].apply(
            lambda x: locale.format_string("%10.0f", x, grouping=True, monetary=True)
        )
        # note : not every numerical column needs a sum line
        # breedte df.shape[1] is het tweede element van de tuple
        # df.shape, het aantal kolomen
        # print(df.shape[1])
        add_separator_line(df, df.shape[1] - 1)
        df = df.fillna("")
        # df = df.astype(str)
        # df = pd.concat([df, df.iloc[-1:]], ignore_index=True)
        # df.at[df.index[-2], df.columns[0]] = ""
        # df.at[df.index[-2], df.columns[1]] = "-" * 7

    return df


def sheet_exists(filename: str, sheet_name: str) -> bool:
    """
    open an Excel file and return a workbook
    file, sheet, table mogelijk in 1 functie samennemen? *kwars
    """
    wb = load_workbook(filename, read_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit("Error: sheet is not in Excel file")
        # The module shall have only one (normal) exit
        # I expect there shall be an error handler (function)
    return True


def add_thousands_separator(df: pd.DataFrame, column: list[str]) -> pd.DataFrame:
    """
    To be added to all numeric columns, not just column[1]
    This will throw an error if there is a non-numeric in column Aantal
    """
    df[column[1]] = df[column[1]].apply(
        lambda x: locale.format_string("%10.0f", x, grouping=True, monetary=True)
    )
    return df


def add_separator_line(df: pd.DataFrame, column_idx: int) -> pd.DataFrame:
    """
    is not yet generic, just works on last column
    add a separator line above the totals
    """
    df = df.astype({df.columns[column_idx]: str})
    df = pd.concat([df, df.iloc[-1:]], ignore_index=True)
    df.at[df.index[-2], df.columns[0]] = ""
    df.at[df.index[-2], df.columns[column_idx]] = "-" * 7
    return df


def unpack_xy(sheet_name_cell_range: str) -> list:
    """
    Given a (table)range this function returns the upper left and lower right
    (table)coordinates in decimal notation.

    Args:
        sheet_name_cell_range (str): the sheetname and the table range

    Example:
        unpack_xy('tabellen!G71:H73')

    Returns:
        a list containing four coordinates x1,x2,y1,y2

    Change (27-12-2023):
        no longer constrained to A:Z and 1:99
    """

    # range_to_tuple() returns a tuple of two elements:
    # sheetname and range. The second element is unpacked.
    xy = range_to_tuple(sheet_name_cell_range)[1]
    return [get_column_letter(xy[0]), xy[1], get_column_letter(xy[2]), xy[3]]
