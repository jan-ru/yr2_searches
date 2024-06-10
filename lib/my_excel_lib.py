""" This module contains various functions dealing with excel files.
    create_table, expand_table, read_table

xls_tables, finds the sheets and tables in an excel file.

Args:
    file_name (str): the filename of the excel data file

Example:
    xls_tables('sia-projecten.xlsx')

Attributes:
    df is the Pandas DataFrame that the excel table is loaded into

To do:
     * generalize the module such that it accepts tables with multiple
        numeric columns maybe later pass a list of dictionary's where
        the dictionary holds {column name : sum=true} or {column name
        : sum = false}
     * include a test check_table_exists -> True/False
     * possibly use argparse as shown in excel2markdown by SpeerSec.

.. _Ref excel2markdown by SpeerSec:
   https://github.com/SpeerSec/excel2markdown
"""

import sys
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def get_tables(excel_file_name: str) -> dict:
    """Get all tables from a given workbook. Returns a dictionary of tables.
    Requires a filename, which includes the path and filename.
    """
    tables = {}

    wb = load_workbook(excel_file_name)
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            # print("table aantal", len(ws.tables))
            tables[tbl.name] = {
                "file": excel_file_name,
                "sheet": ws.title,
                "name": tbl.name,
                "range": tbl.ref,
                "columns": tbl.tableColumns,
            }
            # print(" : " + tbl.displayName)
            # print("   - #cols = %d" % len(tbl.tableColumns))
            # for col in tbl.tableColumns:
            #        print("2")
            # print("     : " + col.name)

    return tables


def column_number_string(n):
    """
    number is converted into a string

    Example:
        column_number_string(1) results in 'a'

    """
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def expand_table():
    """
    from a workbook (wb) and a worksheet (ws) a table is created?
    """
    table_name = "Data"

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    wb = load_workbook(filename="workbook.xlsx")
    ws = wb["inputData"]

    with open("input.csv", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=";")
        for i, row in enumerate(reader):
            for j, cell in enumerate(row):
                if not i == 0:
                    ws.cell(row=i + 1, column=j + 1).value = float(cell)
                else:
                    ws.cell(row=i + 1, column=j + 1).value = cell

                max_ref = [i, j]

    for i, table in enumerate(ws._tables):
        if table.name == table_name:
            table_ref = i

    res_table = Table(
        displayName="Data",
        ref="A1:{}{}".format(column_number_string(max_ref[0]), max_ref[1]),
    )
    res_table.tableStyleInfo = style

    ws._tables[table_ref] = res_table

    wb.save("output.xlsx")


# Creating a table
def create_tabel(table_name: str):
    """
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unque through out defined names
    and all other table name.
    not correct
    """
    if table_name not in ["a", "b"]:
        sys.exit("Error: table not in Excel file")
    wb = Workbook()
    ws = wb.active

    data = [
        ["Apples", 10000, 5000, 8000, 6000],
        ["Pears", 2000, 3000, 4000, 5000],
        ["Bananas", 6000, 6000, 6500, 6000],
        ["Oranges", 500, 300, 200, 700],
    ]

    # add column headings. NB. these must be strings
    ws.append(["Fruit", "2011", "2012", "2013", "2014"])
    for row in data:
        ws.append(row)

    tab = Table(displayName="Table1", ref="A1:E5")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)
    wb.save("table.xlsx")
